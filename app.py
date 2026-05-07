import math
import re
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# =========================
# 页面设置
# =========================
st.set_page_config(page_title="终末地计算工具", layout="wide")

st.markdown("""
<style>
:root{
    --bg:#f5f7fb;
    --card:#ffffff;
    --soft:#eaf3ff;
    --soft-border:#d8e8ff;
    --text:#111827;
    --muted:#6b7280;
    --dark:#1f1f1f;
    --yellow:#ffd900;
}
html, body, [data-testid="stAppViewContainer"]{
    background:var(--bg);
}
.block-container{
    padding-top:4.0rem;
    padding-left:2.2rem;
    padding-right:2.2rem;
    max-width:1500px;
}
h1,h2,h3{
    margin-top:0.35rem;
    margin-bottom:0.7rem;
    line-height:1.35 !important;
    overflow:visible !important;
    color:var(--text);
}
div[data-testid="stVerticalBlock"]{gap:0.65rem;}
.app-title{
    font-size:2.05rem;
    font-weight:900;
    line-height:1.35;
    margin:0.5rem 0 1.0rem 0;
    color:var(--text);
    letter-spacing:.02em;
}
.section-card{
    background:var(--card);
    border:1px solid #e5e7eb;
    border-radius:18px;
    padding:18px 20px;
    margin:0.35rem 0 1.2rem 0;
    box-shadow:0 8px 24px rgba(15,23,42,.05);
}
.info-box{
    background:var(--soft);
    border-radius:14px;
    padding:14px 16px;
    min-height:82px;
    border:1px solid var(--soft-border);
    display:flex;
    flex-direction:column;
    justify-content:center;
}
.info-title{font-size:13px;color:var(--muted);margin-bottom:7px;font-weight:600;}
.info-value{font-size:19px;font-weight:800;color:var(--text);line-height:1.25;}
.weapon-card,.equip-card,.potential-card{
    background:var(--soft);
    padding:14px 16px;
    border-radius:14px;
    margin-bottom:8px;
    min-height:112px;
    border:1px solid var(--soft-border);
    line-height:1.55;
}
.potential-card.locked{
    background:#f3f4f6;
    border-color:#e5e7eb;
    color:#6b7280;
}
.stat-card{
    background:var(--dark);
    color:white;
    padding:12px 16px;
    border-radius:10px;
    margin-bottom:7px;
    display:flex;
    justify-content:space-between;
    align-items:center;
    font-size:18px;
}
.ability-card{
    background:#333;
    color:white;
    padding:15px 18px;
    border-radius:12px;
    margin-bottom:10px;
}
.ability-main{
    background:var(--yellow);
    color:#111;
    padding:15px 18px;
    border-radius:12px;
    margin-bottom:10px;
}
.small-note{color:#555;font-size:13px;line-height:1.65;}
/* Streamlit controls spacing */
div[data-testid="stSidebar"] .block-container{padding-top:2.0rem;}
div[data-testid="stSlider"]{padding-top:.2rem;}
div[data-testid="stSelectbox"]{margin-bottom:.1rem;}
hr.soft-line{border:none;border-top:1px solid #e5e7eb;margin:10px 0 14px;}

/* 伤害计算页面：让输入框和按钮更紧凑 */
div[data-testid="stNumberInput"] input{min-height:2.15rem !important;padding-top:0.25rem !important;padding-bottom:0.25rem !important;}
div[data-testid="stNumberInput"] button{min-height:2.15rem !important;}
div[data-testid="stSelectbox"] div[data-baseweb="select"]{min-height:2.25rem !important;}
div.stButton > button{min-height:2.25rem; padding-top:0.25rem; padding-bottom:0.25rem;}

</style>
""", unsafe_allow_html=True)

st.markdown("<div class='app-title'>终末地计算工具</div>", unsafe_allow_html=True)

BASE_DIR = Path(__file__).resolve().parent
CHARACTER_EXCEL = BASE_DIR / "终末地干员数据.xlsx"
WEAPON_EXCEL = BASE_DIR / "终末地武器数据.xlsx"
EQUIP_EXCEL = BASE_DIR / "终末地装备数据.xlsx"

# =========================
# 读取数据：全部缓存，减少每次滑条白屏时间
# =========================
@st.cache_data(show_spinner=False)
def load_data(character_path: str, weapon_path: str, equip_path: str):
    c = pd.read_excel(character_path, sheet_name="DataSource").dropna(how="all")
    w = pd.read_excel(weapon_path, sheet_name="DataSource").dropna(how="all")
    e = pd.read_excel(equip_path, sheet_name="DataSource").dropna(how="all")
    return c, w, e

@st.cache_data(show_spinner=False)
def load_equipment_formula_map(path: str, sheet_name="终末地装备数据"):
    """装备副词条是 CHOOSE 公式，pandas 经常读不到显示值，所以缓存公式原文。"""
    formula_cols = {
        "①": "词条①精锻数值（含不可见的精确部分）",
        "②": "词条②精锻数值（含不可见的精确部分）",
        "③": "词条③精锻数值（含不可见的精确部分）",
    }
    formula_map = {}
    try:
        wb = load_workbook(path, data_only=False, read_only=True)
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_pos = {idx: headers.index(col) + 1 for idx, col in formula_cols.items() if col in headers}
        for excel_row in range(2, ws.max_row + 1):
            formula_map[excel_row] = {}
            for idx, excel_col in col_pos.items():
                formula_map[excel_row][idx] = ws.cell(excel_row, excel_col).value
    except Exception:
        return {}
    return formula_map


@st.cache_data(show_spinner=False)
def load_character_skill_data(character_path: str, character_name: str):
    """读取某个干员的技能表。
    工作表格式通常为：xxx_技能，包含四个技能块：
    ✦ 【普通攻击】...
    ✦ 【战技】...
    ✦ 【终结技】...
    ✦ 【连携技】...
    每个技能块下方有参数名称/Lv1-Lv9/M1-M3。
    """
    sheet_name = f"{character_name}_技能"
    try:
        xls = pd.ExcelFile(character_path)
        if sheet_name not in xls.sheet_names:
            return {}
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    except Exception:
        return {}

    blocks = []
    first_col = df.iloc[:, 0].tolist()
    for i, value in enumerate(first_col):
        if isinstance(value, str) and "【" in value and "】" in value:
            title = value.strip()
            skill_type = None
            for key in ["普通攻击", "战技", "连携技", "终结技"]:
                if f"【{key}】" in title:
                    skill_type = key
                    break
            if skill_type:
                blocks.append((i, skill_type, title))

    skills = {}
    for idx, (start_row, skill_type, title) in enumerate(blocks):
        end_row = blocks[idx + 1][0] if idx + 1 < len(blocks) else len(df)
        desc = ""
        if start_row + 1 < len(df):
            desc_val = df.iloc[start_row + 1, 0]
            if isinstance(desc_val, str):
                desc = desc_val.strip()

        header_row = None
        for r in range(start_row + 1, end_row):
            v = df.iloc[r, 0]
            if isinstance(v, str) and "参数名称" in v:
                header_row = r
                break

        params = []
        if header_row is not None:
            headers = [str(x).strip() if not pd.isna(x) else "" for x in df.iloc[header_row].tolist()]
            for r in range(header_row + 1, end_row):
                name = df.iloc[r, 0]
                if pd.isna(name):
                    continue
                name = str(name).strip()
                if not name:
                    continue
                row_data = {"参数名称": name}
                for c, h in enumerate(headers):
                    if h:
                        val = df.iloc[r, c]
                        if not pd.isna(val):
                            row_data[h] = val
                params.append(row_data)

        skills[skill_type] = {
            "标题": title,
            "描述": desc,
            "参数": params,
        }

    return skills


def skill_level_to_col(level_name: str):
    mapping = {
        "专1": "M1",
        "专2": "M2",
        "专3": "M3",
    }
    return mapping.get(level_name, level_name)


def show_skill_card(skill_name, skill_data, level_name, key_prefix):
    col_name = skill_level_to_col(level_name)

    if not skill_data:
        st.markdown(
            f"<div class='weapon-card'><b>{skill_name}</b><br><br>当前干员没有找到技能表数据。</div>",
            unsafe_allow_html=True
        )
        return

    title = skill_data.get("标题", skill_name)
    desc = skill_data.get("描述", "")
    params = skill_data.get("参数", [])

    lines = []
    for row in params:
        param_name = row.get("参数名称", "")
        value = row.get(col_name, "")
        if str(value).strip() and str(value).strip() != "nan":
            lines.append(f"<tr><td style='padding:4px 0;border-top:1px solid #d8e8ff;'>{param_name}</td><td style='padding:4px 0;border-top:1px solid #d8e8ff;text-align:right;'><b>{value}</b></td></tr>")

    table_html = ""
    if lines:
        table_html = (
            "<table style='width:100%;border-collapse:collapse;margin-top:10px;font-size:14px;'>"
            + "".join(lines)
            + "</table>"
        )

    st.markdown(
        f"""
        <div class='weapon-card' style='min-height:180px;'>
            <b>{skill_name}｜{level_name}</b><br>
            <span style='font-size:13px;color:#555;'>{title}</span>
            <div style='margin-top:8px;font-size:14px;line-height:1.65;'>{desc}</div>
            {table_html}
        </div>
        """,
        unsafe_allow_html=True
    )

def parse_multiplier_value(value):
    """
    从技能参数里提取倍率数字。
    支持：
    120%
    120.5%
    攻击力120%
    120%攻击力
    """
    text = clean_text(value)
    if not text or text == "无":
        return 0.0

    m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*[%％]", text)
    if m:
        return float(m.group(1))

    return 0.0


def get_skill_multiplier_summary(character_name, level_name="Lv9"):
    """
    返回某个干员的技能倍率汇总。
    普攻：1-5段合并
    处决：单独
    战技 / 连携技 / 终结技：分别汇总
    """
    skill_data_map = load_character_skill_data(str(CHARACTER_EXCEL), character_name)
    col_name = skill_level_to_col(level_name)

    result = {
        "普通攻击": 0.0,
        "处决": 0.0,
        "战技": 0.0,
        "连携技": 0.0,
        "终结技": 0.0,
    }

    for skill_name, skill_data in skill_data_map.items():
        params = skill_data.get("参数", [])

        for row in params:
            param_name = clean_text(row.get("参数名称", ""))
            value = row.get(col_name, "")
            mult = parse_multiplier_value(value)

            if mult <= 0:
                continue

            if skill_name == "普通攻击":
                # 普攻1-5段合并
                if any(x in param_name for x in ["一段", "二段", "三段", "四段", "五段", "第1段", "第2段", "第3段", "第4段", "第5段", "1段", "2段", "3段", "4段", "5段"]):
                    result["普通攻击"] += mult

                # 处决单独算
                elif "处决" in param_name:
                    result["处决"] += mult

            elif skill_name in ["战技", "连携技", "终结技"]:
                # 只要参数里是倍率，就先合计
                result[skill_name] += mult

    return result

def get_skill_multiplier_detail_by_levels(character_name, skill_level_map):
    skill_data_map = load_character_skill_data(str(CHARACTER_EXCEL), character_name)

    result = {
        "普通攻击": [],
        "处决": [],
        "战技": [],
        "连携技": [],
        "终结技": [],
    }

    for skill_name, skill_data in skill_data_map.items():
        level_name = skill_level_map.get(skill_name, "Lv9")
        col_name = skill_level_to_col(level_name)

        params = skill_data.get("参数", [])

        for row in params:
            param_name = clean_text(row.get("参数名称", ""))
            value = row.get(col_name, "")
            mult = parse_multiplier_value(value)

            if mult <= 0:
                continue

            item = {
                "名称": param_name,
                "倍率": mult,
                "原始值": value,
                "技能等级": level_name,
            }

            if skill_name == "普通攻击":
                if "处决" in param_name:
                    result["处决"].append(item)
                else:
                    result["普通攻击"].append(item)

            elif skill_name in ["战技", "连携技", "终结技"]:
                result[skill_name].append(item)

    return result

for path in [CHARACTER_EXCEL, WEAPON_EXCEL, EQUIP_EXCEL]:
    if not path.exists():
        st.error(f"找不到文件：{path.name}。请把 Excel 和本 py 文件放在同一个文件夹。")
        st.stop()

characters_df, weapon_df, equip_df = load_data(str(CHARACTER_EXCEL), str(WEAPON_EXCEL), str(EQUIP_EXCEL))
equip_formula_map = {}

# =========================
# 基础工具函数
# =========================
def safe_get(row, col, default=""):
    try:
        value = row[col]
        if pd.isna(value):
            return default
        return value
    except Exception:
        return default


def to_num(value, default=0.0):
    try:
        if value is None or pd.isna(value):
            return default
        if isinstance(value, str):
            value = value.replace("%", "").replace("％", "").replace("＋", "+").strip()
        return float(value)
    except Exception:
        return default


def floor_num(value):
    return int(math.floor(to_num(value, 0)))


def ceil_num(value):
    return int(math.ceil(to_num(value, 0)))


def clean_text(text):
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return ""
    return str(text).replace("\t", "").replace(" ", "").replace("＋", "+").replace("％", "%").strip()


def strip_title(text):
    # 去掉【xxx Lv.9】标题，避免 Lv.9 被当成词条数值
    text = clean_text(text)
    return re.sub(r"^【[^】]*】", "", text)


def pct_fraction(value, default=0.0):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    if isinstance(value, str):
        raw = value.strip().replace("％", "%")
        num = to_num(raw, default * 100 if "%" in raw else default)
        if "%" in raw:
            return num / 100
        return num / 100 if abs(num) > 1 else num
    num = to_num(value, default)
    return num / 100 if abs(num) > 1 else num


def format_percent(value):
    return f"{to_num(value, 0) * 100:.1f}%"


def attr_res_from_ability(value):
    # 和游戏显示更接近的递减抗性公式：敏捷160≈14，敏捷662≈40
    v = to_num(value, 0)
    return int(round(100 * v / (1000 + v))) if v > 0 else 0


def get_level_row(df, name_col, name, level_col, level):
    temp = df[df[name_col] == name].copy()
    if len(temp) == 0:
        return None
    temp["__level_num"] = pd.to_numeric(temp[level_col], errors="coerce")
    exact = temp[temp["__level_num"] == level]
    if len(exact) > 0:
        return exact.iloc[0]
    lower = temp[temp["__level_num"] <= level].sort_values("__level_num")
    if len(lower) > 0:
        return lower.iloc[-1]
    return temp.iloc[0]


def collect_texts(*items):
    out = []
    for item in items:
        if item is None:
            continue
        if isinstance(item, (list, tuple)):
            for x in item:
                t = clean_text(x)
                if t and t != "无":
                    out.append(t)
        else:
            t = clean_text(item)
            if t and t != "无":
                out.append(t)
    return out

# =========================
# 词条解析：固定值和百分比彻底分开
# =========================
def parse_key_values(text, keys):
    """返回 [(key, value, is_percent)]。
    支持：力量+15、敏捷能力值提升20、攻击力+39%、主能力值+156、副能力+14.0%。
    一段文本里有多个属性时逐个解析，不会因为后面有百分号把前面的固定值误判成百分比。
    """
    text = strip_title(text)
    if not text:
        return []
    keys = sorted([k for k in keys if k], key=len, reverse=True)
    results = []
    used = []
    for key in keys:
        pattern = re.compile(
            re.escape(key) +
            r"(?:能力值)?(?:提升|提高|增加|加成|效率|数值)?\s*[:：]?\s*[+＋]?\s*" +
            r"([+-]?\d+(?:\.\d+)?)(\s*[%％])?"
        )
        for m in pattern.finditer(text):
            span = m.span()
            if any(not (span[1] <= u[0] or span[0] >= u[1]) for u in used):
                continue
            used.append(span)
            results.append((key, float(m.group(1)), bool(m.group(2))))
    return results


def stat_flat_from_text(text, keys):
    return sum(v for _, v, is_pct in parse_key_values(text, keys) if not is_pct)


def stat_percent_from_text(text, keys):
    return sum(v / 100 for _, v, is_pct in parse_key_values(text, keys) if is_pct)


def stat_any_number(text, keys):
    vals = parse_key_values(text, keys)
    return sum(v for _, v, _ in vals)


def text_has_any(text, words):
    text = clean_text(text)
    return any(w in text for w in words)


def ability_keys_for(ability_name, main_attr, sub_attr):
    keys = [f"{ability_name}能力值", ability_name]
    if ability_name == main_attr:
        keys.extend(["主能力值", "主能力", "主属性"])
    if ability_name == sub_attr:
        keys.extend(["副能力值", "副能力", "副属性"])
    return keys


def ability_flat_from_text(text, ability_name, main_attr, sub_attr):
    return stat_flat_from_text(text, ability_keys_for(ability_name, main_attr, sub_attr))


def ability_percent_from_text(text, ability_name, main_attr, sub_attr):
    return stat_percent_from_text(text, ability_keys_for(ability_name, main_attr, sub_attr))


def calc_ability(base_value, ability_name, texts, main_attr, sub_attr):
    base = to_num(base_value, 0)

    # 单项能力：力量/敏捷/智识/意志/主能力/副能力
    flat = sum(
        ability_flat_from_text(t, ability_name, main_attr, sub_attr)
        for t in texts
    )

    pct = sum(
        ability_percent_from_text(t, ability_name, main_attr, sub_attr)
        for t in texts
    )

    # 全能力：只在这里统一加一次，不要写进 ability_flat_from_text
    all_flat = sum(
        stat_flat_from_text(t, ["全能力值", "全能力"])
        for t in texts
    )

    all_pct = 0

    flat += all_flat
    pct += all_pct

    final = floor_num((base + flat) * (1 + pct))
    return final, flat, pct


def attack_percent_from_text(text):
    if text is None or pd.isna(text):
        return 0.0

    text = strip_title(str(text))

    # 按句子拆开，避免“前半句常驻，后半句条件”被整条排除
    parts = re.split(r"[。；;]", text)

    total = 0.0

    ban_words = [
        "造成",
        "命中",
        "持续",
        "每层",
        "触发",
        "额外",
        "战技",
        "终结技",
        "连携技",
        "普通攻击",
        "伤害",
        "提升效果",
        "攻击倍率",
        "敌人",
        "目标",
        "场上每有",
        "状态",
    ]

    for part in parts:
        part = part.strip()
        if not part:
            continue

        # 只排除当前这一小句，不排除整条词条
        if any(word in part for word in ban_words):
            continue

        m = re.search(
            r"攻击力\s*[+＋]\s*([0-9]+(?:\.[0-9]+)?)\s*[%％]",
            part
        )

        if m:
            total += float(m.group(1)) / 100

    return total
    # 只识别真正面板攻击%
    m = re.search(
        r"攻击力\s*[+＋]\s*([0-9]+(?:\.[0-9]+)?)\s*[%％]",
        text
    )

    if m:
        return float(m.group(1)) / 100

    return 0.0


def fixed_attack_from_text(text):
    return stat_flat_from_text(text, ["固定攻击力", "攻击力固定"])


def panel_bonus_percent(text, keys):
    vals = parse_key_values(text, keys)
    if not vals:
        return 0.0
    # 面板百分比列：公式里没 % 时也按百分比数值理解
    return vals[0][1] / 100


def physical_damage_bonus_from_text(text, source=""):
    text = clean_text(text)
    if not text or text == "无":
        return 0.0

    # 明确面板列，优先算
    v = panel_bonus_percent(text, ["物理伤害加成"])
    if v:
        return v

    # 武器词条“物理伤害+xx%”是常驻面板
    if source == "weapon":
        m = re.search(r"物理伤害\s*[+＋]\s*([0-9]+(?:\.[0-9]+)?)\s*[%％]", strip_title(text))
        return float(m.group(1)) / 100 if m else 0.0

    # 潜能里“造成的物理伤害+8%”通常是常驻；带持续/命中/期间等条件的不算
    if source == "potential":
        if text_has_any(text, ["持续", "命中", "触发", "如果", "若", "当", "期间", "下一次", "下次", "受到的物理伤害", "额外造成", "造成一次"]):
            return 0.0
        m = re.search(r"(?:造成的)?物理伤害\s*[+＋]\s*([0-9]+(?:\.[0-9]+)?)\s*[%％]", strip_title(text))
        return float(m.group(1)) / 100 if m else 0.0

    return 0.0


def collect_physical_bonus(weapon_texts, potential_texts, equip_texts=None, set_texts=None):
    items = []
    for source, texts in [("weapon", weapon_texts), ("potential", potential_texts), ("equip", equip_texts or []), ("set", set_texts or [])]:
        for t in texts:
            v = physical_damage_bonus_from_text(t, source)
            if abs(v) > 1e-12:
                items.append((source, clean_text(t), v))
    return items

# =========================
# 武器 / 装备函数
# =========================
def get_weapon_affix(row, affix_index, affix_level):
    source_col = f"__weapon_skill_{'①②③'.index(affix_index) + 1}_lv{affix_level}"
    direct_col = f"武器词条{affix_index}"
    if row is not None and source_col in row.index:
        value = safe_get(row, source_col, "无")
        if clean_text(value) and clean_text(value) != "无":
            return value
    return safe_get(row, direct_col, "无") if row is not None else "无"


def eval_choose_formula(value, level=0):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "无"
    text = str(value).strip()
    if not text:
        return "无"
    if not text.startswith("=") or "CHOOSE" not in text.upper():
        return text
    options = re.findall(r'"([^"]*)"', text)
    if not options:
        return "无"
    lv = max(0, min(int(to_num(level, 0)), len(options) - 1))
    return options[lv]


def excel_row_number(row):
    try:
        if "index" in row.index:
            return int(row["index"]) + 2
    except Exception:
        pass
    try:
        return int(row.name) + 2
    except Exception:
        return None


def equip_affix_value(row, idx, refine_level=0):
    if row is None:
        return "无"
    excel_row = excel_row_number(row)
    raw = None
    if excel_row in equip_formula_map:
        raw = equip_formula_map.get(excel_row, {}).get(idx)
    if raw is None or raw == "" or (isinstance(raw, float) and pd.isna(raw)):
        raw = safe_get(row, f"词条{idx}精锻数值（含不可见的精确部分）", "无")
    return eval_choose_formula(raw, refine_level)


def get_equip_texts(row, refine_levels=None):
    if row is None:
        return []
    refine_levels = refine_levels or {"①": 0, "②": 0, "③": 0}
    return [
        safe_get(row, "主词条数值", "无"),
        equip_affix_value(row, "①", refine_levels.get("①", 0)),
        equip_affix_value(row, "②", refine_levels.get("②", 0)),
        equip_affix_value(row, "③", refine_levels.get("③", 0)),
    ]

# =========================
# 显示函数
# =========================
def info_box(title, value):
    st.markdown(f"""
    <div class='info-box'>
        <div class='info-title'>{title}</div>
        <div class='info-value'>{value}</div>
    </div>
    """, unsafe_allow_html=True)


def show_stat(name, value):
    st.markdown(f"<div class='stat-card'><span>{name}</span><b>{value}</b></div>", unsafe_allow_html=True)


def show_ability(name, value, detail="", main=False):
    css = "ability-main" if main else "ability-card"
    st.markdown(f"""
    <div class='{css}'>
        <div style='font-size:20px;font-weight:700;'>{name}</div>
        <div style='font-size:32px;font-weight:700;'>{value}</div>
        <div style='font-size:13px;opacity:.75;'>{detail}</div>
    </div>
    """, unsafe_allow_html=True)

# =========================
# 全局编队数据：两个页面都要用
# =========================
character_names = characters_df["干员名称"].dropna().drop_duplicates().tolist()
character_names = [x for x in character_names if x != "管理员"]

slot_labels = ["1号位", "2号位", "3号位", "4号位"]

TEAM_WIDGET_KEYS = [
    "selected_character",
    "character_level",
    "potential_level",
    "trust_level",
    "selected_weapon_display",
    "selected_weapon_level",
    "affix1_level",
    "affix2_level",
    "affix3_level",
    "enable_equipment",
    "armor_select", "armor_a1", "armor_a2", "armor_a3",
    "hand_select", "hand_a1", "hand_a2", "hand_a3",
    "acc1_select", "acc1_a1", "acc1_a2", "acc1_a3",
    "acc2_select", "acc2_a1", "acc2_a2", "acc2_a3",
    "skill_level_普通攻击",
    "skill_level_战技",
    "skill_level_连携技",
    "skill_level_终结技",
]


def make_default_profile():
    return {
        "selected_character": character_names[0] if character_names else "",
        "potential_level": 0,
        "trust_level": 4,
        "enable_equipment": False,
        "affix1_level": 9,
        "affix2_level": 9,
        "affix3_level": 9,
        "armor_select": "不装备",
        "hand_select": "不装备",
        "acc1_select": "不装备",
        "acc2_select": "不装备",
        "armor_a1": 3, "armor_a2": 3, "armor_a3": 3,
        "hand_a1": 3, "hand_a2": 3, "hand_a3": 3,
        "acc1_a1": 3, "acc1_a2": 3, "acc1_a3": 3,
        "acc2_a1": 3, "acc2_a2": 3, "acc2_a3": 3,
        "skill_level_普通攻击": "Lv9",
        "skill_level_战技": "Lv9",
        "skill_level_连携技": "Lv9",
        "skill_level_终结技": "Lv9",
    }


def save_current_slot(slot_index):
    profile = st.session_state.team_profiles[slot_index]
    for key in TEAM_WIDGET_KEYS:
        if key in st.session_state:
            profile[key] = st.session_state[key]

def save_before_page_change():
    if "team_profiles" in st.session_state and "active_slot" in st.session_state:
        save_current_slot(st.session_state.active_slot)

def restore_slot(slot_index):
    profile = st.session_state.team_profiles[slot_index]
    if profile.get("selected_character") not in character_names:
        profile["selected_character"] = character_names[0] if character_names else ""
    for key, value in profile.items():
        st.session_state[key] = value

st.sidebar.header("功能页面")
page_mode = st.sidebar.radio(
    "选择页面",
    ["编队设置", "伤害计算"],
    key="page_mode",
)

st.sidebar.markdown("---")

if "team_profiles" not in st.session_state:
    st.session_state.team_profiles = [make_default_profile() for _ in range(4)]

if "team_panel_results" not in st.session_state:
    st.session_state.team_panel_results = {}

if "team_medicine_choices" not in st.session_state:
    st.session_state.team_medicine_choices = {
        "1号位": "不使用",
        "2号位": "不使用",
        "3号位": "不使用",
        "4号位": "不使用",
    }

if "damage_axis_rows" not in st.session_state:
    st.session_state.damage_axis_rows = []

if "damage_axis_editor_version" not in st.session_state:
    st.session_state.damage_axis_editor_version = 0

if page_mode == "编队设置":
    st.sidebar.header("四人编队")

    # 关键修复：Streamlit 在切到“伤害计算”时不会渲染编队控件，
    # 返回时同名 widget key 可能被清理，必须在创建 radio/selectbox 前恢复当前槽位。
    returning_from_other_page = st.session_state.get("_last_page") != "编队设置"
    if returning_from_other_page:
        active_slot_for_restore = st.session_state.get("active_slot", 0)
        active_slot_for_restore = max(0, min(int(active_slot_for_restore), len(slot_labels) - 1))
        st.session_state["active_slot_radio"] = slot_labels[active_slot_for_restore]
        restore_slot(active_slot_for_restore)

    active_slot_label = st.sidebar.radio(
        "当前编辑",
        slot_labels,
        index=st.session_state.get("active_slot", 0),
        key="active_slot_radio"
    )

    active_slot = slot_labels.index(active_slot_label)

    previous_slot = st.session_state.get("active_slot")

    if previous_slot is None:
        st.session_state.active_slot = active_slot
        restore_slot(active_slot)
    elif previous_slot != active_slot:
        save_current_slot(previous_slot)
        st.session_state.active_slot = active_slot
        restore_slot(active_slot)

    st.sidebar.markdown("---")
    st.sidebar.header("角色选择")

    selected_character = st.sidebar.selectbox(
        "选择干员",
        character_names,
        key="selected_character",
    )

    st.sidebar.caption("切换 1–4 号位时会自动保存当前配置，并恢复对应槽位配置。")
    # =========================
    # 基础信息：等级 / 潜能 / 信赖放在一个清晰的角色信息卡片里
    # =========================
    st.markdown("## 基础信息")
    char_rows = characters_df[characters_df["干员名称"] == selected_character]
    level_values = pd.to_numeric(char_rows["等级"], errors="coerce").dropna()
    level_min = int(level_values.min())
    level_max = min(int(level_values.max()), 90)
    if "character_level" not in st.session_state or st.session_state.character_level < level_min or st.session_state.character_level > level_max:
        st.session_state.character_level = level_max

    with st.container():
        st.markdown("<div class='section-card'>", unsafe_allow_html=True)

        ctrl_cols = st.columns([1.35, 1.35, 1.0, 1.0])
        with ctrl_cols[0]:
            info_box("当前干员", selected_character)
        with ctrl_cols[1]:
            character_level = st.slider(
                "等级",
                level_min,
                level_max,
                level_max,
                key="character_level",
                help="调整干员等级，面板基础值会随等级变化。"
            )
        with ctrl_cols[2]:
            potential_level = st.selectbox(
                "潜能",
                [0, 1, 2, 3, 4, 5],
                index=0,
                key="potential_level"
            )
        with ctrl_cols[3]:

            trust_options = [0, 1, 2, 3, 4]

            if "trust_level" not in st.session_state:
                st.session_state["trust_level"] = 4

            trust_level = st.selectbox(
                "信赖节点",
                trust_options,
                index=trust_options.index(st.session_state.get("trust_level", 4)),
                key="trust_level"
            )
        character_data = get_level_row(characters_df, "干员名称", selected_character, "等级", character_level)
        main_attr = clean_text(safe_get(character_data, "主属性", ""))
        sub_attr = clean_text(safe_get(character_data, "副属性", ""))

        st.markdown("<hr class='soft-line'>", unsafe_allow_html=True)

        detail_cols = st.columns(5)
        with detail_cols[0]:
            info_box("职业", safe_get(character_data, "职业", ""))
        with detail_cols[1]:
            info_box("干员属性", safe_get(character_data, "干员属性", ""))
        with detail_cols[2]:
            info_box("主 / 副属性", f"{main_attr} / {sub_attr}")
        with detail_cols[3]:
            info_box("武器类型", safe_get(character_data, "武器类型", ""))
        with detail_cols[4]:
            info_box("稀有度", safe_get(character_data, "稀有度", ""))

        st.markdown("<hr class='soft-line'>", unsafe_allow_html=True)
        st.markdown("#### 潜能效果")

        p_cols = st.columns(5)
        potential_texts = []
        for i in range(1, 6):
            col_name = f"潜能天赋{'①②③④⑤'[i-1]}"
            text = safe_get(character_data, col_name, "无")
            unlocked = i <= potential_level
            if unlocked and clean_text(text) and clean_text(text) != "无":
                potential_texts.append(text)
            with p_cols[i - 1]:
                label = text if unlocked else "未解锁"
                locked_class = "" if unlocked else " locked"
                st.markdown(
                    f"<div class='potential-card{locked_class}'><b>潜能天赋{i}</b><br><br>{label}</div>",
                    unsafe_allow_html=True
                )

        st.markdown("</div>", unsafe_allow_html=True)

        trust_texts = []

        trust_data = char_rows.copy()
        trust_data["__level_num"] = pd.to_numeric(trust_data["等级"], errors="coerce")
        trust_data = trust_data.sort_values("__level_num").iloc[-1]

        for i in range(1, trust_level + 1):
            col_name = f"信赖天赋节点{'①②③④'[i-1]}"
            text = safe_get(trust_data, col_name, "无")
            if clean_text(text) and clean_text(text) != "无":
                trust_texts.append(text)

    # =========================
    # 武器系统：基础攻击在上面，每个词条下面对应等级
    # =========================
    st.markdown("## 武器系统")
    weapon_type = safe_get(character_data, "武器类型", "")
    filtered_weapon_df = weapon_df[weapon_df["武器类型"] == weapon_type].copy().dropna(subset=["武器名称"])
    if len(filtered_weapon_df) == 0:
        st.error(f"没有找到武器类型为【{weapon_type}】的武器。")
        st.stop()

    weapon_display_map = {}
    for _, row in filtered_weapon_df.drop_duplicates(subset=["武器名称", "稀有度"]).iterrows():
        weapon_display_map[f"{safe_get(row, '武器名称')} {safe_get(row, '稀有度', '')}"] = safe_get(row, "武器名称")

    w_top1, w_top2, w_top3 = st.columns([2.0, 1.4, 1.4])
    weapon_display_options = list(weapon_display_map.keys())
    if "selected_weapon_display" not in st.session_state or st.session_state.selected_weapon_display not in weapon_display_options:
        st.session_state.selected_weapon_display = weapon_display_options[0]
    with w_top1:
        selected_weapon_display = st.selectbox(
            "选择武器",
            weapon_display_options,
            key="selected_weapon_display",
        )
    selected_weapon = weapon_display_map[selected_weapon_display]
    weapon_level_df = filtered_weapon_df[filtered_weapon_df["武器名称"] == selected_weapon].copy()
    weapon_levels = pd.to_numeric(weapon_level_df["等级"], errors="coerce").dropna().sort_values().unique().tolist()
    weapon_level_min = int(min(weapon_levels))
    weapon_level_max = int(max(weapon_levels))
    if "selected_weapon_level" not in st.session_state or st.session_state.selected_weapon_level < weapon_level_min or st.session_state.selected_weapon_level > weapon_level_max:
        st.session_state.selected_weapon_level = weapon_level_max
    with w_top2:
        selected_weapon_level = st.slider("武器等级", weapon_level_min, weapon_level_max, weapon_level_max, key="selected_weapon_level")
    weapon_data = get_level_row(weapon_level_df, "武器名称", selected_weapon, "等级", selected_weapon_level)
    weapon_base_atk = to_num(safe_get(weapon_data, "基础攻击力", 0), 0)
    with w_top3:
        info_box("基础攻击力", f"{weapon_base_atk:g}")

    wa1, wa2, wa3 = st.columns(3)
    with wa1:
        affix1_level = st.slider("词条①等级", 1, 9, 9, key="affix1_level")
        affix1 = get_weapon_affix(weapon_data, "①", affix1_level)
        st.markdown(f"<div class='weapon-card'><b>词条① Lv.{affix1_level}</b><br><br>{affix1}</div>", unsafe_allow_html=True)
    with wa2:
        affix2_level = st.slider("词条②等级", 1, 9, 9, key="affix2_level")
        affix2 = get_weapon_affix(weapon_data, "②", affix2_level)
        st.markdown(f"<div class='weapon-card'><b>词条② Lv.{affix2_level}</b><br><br>{affix2}</div>", unsafe_allow_html=True)
    with wa3:
        affix3_level = st.slider("词条③等级", 1, 9, 9, key="affix3_level")
        affix3 = get_weapon_affix(weapon_data, "③", affix3_level)
        st.markdown(f"<div class='weapon-card'><b>词条③ Lv.{affix3_level}</b><br><br>{affix3}</div>", unsafe_allow_html=True)

    weapon_texts = collect_texts(affix1, affix2, affix3)

    # =========================
    # 装备系统：默认关闭，先保证无装备面板速度和数值
    # =========================
    st.markdown("## 装备系统")
    enable_equipment = st.checkbox("启用装备系统", value=False, key="enable_equipment")
    equip_rows = []
    equip_texts = []
    set_effect_texts = []

    def short_text(value, max_len=20):
        t = clean_text(value)
        if not t or t == "无":
            return "无"
        return t if len(t) <= max_len else t[:max_len] + "…"


    def equip_label(row):
        name = safe_get(row, "装备名称", "")
        lv = safe_get(row, "等级", "")
        part = safe_get(row, "部位", "")
        suit = safe_get(row, "套组名称", "")
        main = short_text(safe_get(row, "主词条数值", "无"), 16)
        a1 = short_text(equip_affix_value(row, "①", 0), 16)
        a2 = short_text(equip_affix_value(row, "②", 0), 16)
        a3 = short_text(equip_affix_value(row, "③", 0), 16)
        return f"{name}｜Lv.{lv}｜{part}｜{suit}｜主:{main}｜①:{a1}｜②:{a2}｜③:{a3}｜行{int(row['index'])+2 if 'index' in row.index else ''}"


    def select_equip(slot_name, part, slot_key):
        part_df = equip_df[equip_df["部位"] == part].dropna(subset=["装备名称"]).copy()
        options = ["不装备"]
        row_map = {}
        for _, row in part_df.reset_index(drop=False).iterrows():
            lab = equip_label(row)
            options.append(lab)
            row_map[lab] = row
        select_key = f"{slot_key}_select"
        if select_key not in st.session_state or st.session_state[select_key] not in options:
            st.session_state[select_key] = "不装备"
        picked = st.selectbox(slot_name, options, key=select_key)
        if picked == "不装备":
            return None, {"①": 0, "②": 0, "③": 0}
        row = row_map[picked]
        st.markdown(f"<div class='equip-card'><b>{safe_get(row, '装备名称', '')}</b><br>主词条：{safe_get(row, '主词条数值', '无')}</div>", unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            lv1 = st.select_slider(f"{slot_name}①", options=[0, 1, 2, 3], value=3, key=f"{slot_key}_a1")
        with c2:
            lv2 = st.select_slider(f"{slot_name}②", options=[0, 1, 2, 3], value=3, key=f"{slot_key}_a2")
        with c3:
            lv3 = st.select_slider(f"{slot_name}③", options=[0, 1, 2, 3], value=3, key=f"{slot_key}_a3")
        st.caption(f"① {equip_affix_value(row, '①', lv1)}")
        st.caption(f"② {equip_affix_value(row, '②', lv2)}")
        st.caption(f"③ {equip_affix_value(row, '③', lv3)}")
        return row, {"①": lv1, "②": lv2, "③": lv3}

    if enable_equipment:
        equip_formula_map = load_equipment_formula_map(str(EQUIP_EXCEL))
        e1, e2, e3, e4 = st.columns(4)
        with e1:
            armor_row, armor_ref = select_equip("护甲", "护甲", "armor")
        with e2:
            hand_row, hand_ref = select_equip("护手", "护手", "hand")
        with e3:
            acc1_row, acc1_ref = select_equip("配件1", "配件", "acc1")
        with e4:
            acc2_row, acc2_ref = select_equip("配件2", "配件", "acc2")
        equip_rows = [armor_row, hand_row, acc1_row, acc2_row]
        equip_refs = [armor_ref, hand_ref, acc1_ref, acc2_ref]
        for row, ref in zip(equip_rows, equip_refs):
            equip_texts.extend(get_equip_texts(row, ref))
        set_names = [safe_get(r, "套组名称", "") for r in equip_rows if r is not None]
        for suit in set(set_names):
            if suit and set_names.count(suit) >= 3:
                first = next(r for r in equip_rows if r is not None and safe_get(r, "套组名称", "") == suit)
                set_effect_texts.append(safe_get(first, "套组技能描述", ""))
        with st.expander("查看装备词条 / 套装效果"):
            st.write("装备词条：", [clean_text(x) for x in equip_texts if clean_text(x) and clean_text(x) != "无"])
            st.write("套装效果：", [clean_text(x) for x in set_effect_texts if clean_text(x) and clean_text(x) != "无"])
    else:
        st.info("装备系统当前未启用：下面面板只计算干员基础值、信赖/潜能、武器白值和武器词条。")


    # =========================
    # 技能系统：四个技能，等级 Lv1-Lv9 / 专1-专3
    # =========================
    st.markdown("## 技能系统")
    skill_data_map = load_character_skill_data(str(CHARACTER_EXCEL), selected_character)
    skill_level_options = [f"Lv{i}" for i in range(1, 10)] + ["专1", "专2", "专3"]

    skill_ctrl_cols = st.columns(4)
    skill_levels = {}
    for col, skill_name in zip(skill_ctrl_cols, ["普通攻击", "战技", "连携技", "终结技"]):
        with col:
            skill_levels[skill_name] = st.selectbox(
                f"{skill_name}等级",
                skill_level_options,
                index=8,
                key=f"skill_level_{skill_name}"
            )

    skill_card_cols = st.columns(4)
    for col, skill_name in zip(skill_card_cols, ["普通攻击", "战技", "连携技", "终结技"]):
        with col:
            show_skill_card(
                skill_name,
                skill_data_map.get(skill_name, {}),
                skill_levels[skill_name],
                f"skill_{skill_name}"
            )

    # =========================
    # 干员面板：先保证无装备/只带武器能力值正确
    # =========================
    st.markdown("## 干员面板")

    all_texts = collect_texts(weapon_texts, trust_texts, potential_texts, equip_texts, set_effect_texts)

    base_strength = safe_get(character_data, "力量", 0)
    base_agility = safe_get(character_data, "敏捷", 0)
    base_intellect = safe_get(character_data, "智识", 0)
    base_will = safe_get(character_data, "意志", 0)

    strength, str_flat, str_pct = calc_ability(base_strength, "力量", all_texts, main_attr, sub_attr)
    agility, agi_flat, agi_pct = calc_ability(base_agility, "敏捷", all_texts, main_attr, sub_attr)
    intellect, int_flat, int_pct = calc_ability(base_intellect, "智识", all_texts, main_attr, sub_attr)
    will, will_flat, will_pct = calc_ability(base_will, "意志", all_texts, main_attr, sub_attr)

    ability_map = {"力量": strength, "敏捷": agility, "智识": intellect, "意志": will}
    main_total = ability_map.get(main_attr, 0)
    sub_total = ability_map.get(sub_attr, 0)

    # 游戏攻击力乘区：能力值加成主要由主/副属性提供。
    # 用 1.5% / 0.3% 可以对齐“只带武器无装备”截图中敏捷差值带来的攻击差。
    ability_bonus_main = main_total * 0.005
    ability_bonus_sub = sub_total * 0.002
    ability_bonus = ability_bonus_main + ability_bonus_sub

    char_atk = to_num(safe_get(character_data, "攻击力", 0), 0)

    # 基础攻击力
    base_atk_total = char_atk + weapon_base_atk

    # 攻击%
    attack_percent = sum(
        attack_percent_from_text(t) for t in all_texts
    )

    # 固定攻击力
    fixed_attack = sum(
        fixed_attack_from_text(t) for t in all_texts
    )

    # 百分比攻击加成值
    percent_attack_bonus = base_atk_total * attack_percent

    # 基础总值
    base_total_attack = (
        base_atk_total
        + fixed_attack
        + percent_attack_bonus
    )

    # 最终攻击力
    atk = floor_num(
        base_total_attack * (1 + ability_bonus)
    )

    debug_df = pd.DataFrame([
        ["干员攻击力", round(char_atk, 2)],
        ["武器攻击力", round(weapon_base_atk, 2)],
        ["基础攻击力", round(base_atk_total, 2)],
        ["攻击力%", f"{round(attack_percent * 100, 1)}%"],
        ["百分比加成值", round(percent_attack_bonus, 2)],
        ["固定攻击力", round(fixed_attack, 2)],
        ["基础总值", round(base_total_attack, 2)],
        ["主属性加成", f"{round(ability_bonus_main * 100, 1)}%"],
        ["副属性加成", f"{round(ability_bonus_sub * 100, 1)}%"],
        ["总能力乘区", f"{round(ability_bonus * 100, 1)}%"],
        ["最终攻击力", atk]
    ], columns=["项目", "数值"])

    with st.expander("攻击力来源 / 调试", expanded=False):
        st.table(debug_df)

    def related_ability_text(text, attr_name, main_attr, sub_attr):
        text = clean_text(text)
        if not text or text == "无":
            return False

        keywords = [attr_name, f"{attr_name}能力值"]

        if attr_name == main_attr:
            keywords += ["主能力值", "主能力", "主属性"]

        if attr_name == sub_attr:
            keywords += ["副能力值", "副能力", "副属性"]

        keywords += ["全能力值", "全能力"]

        return any(k in text for k in keywords)


    ability_source_rows = []

    for attr_name in [main_attr, sub_attr]:
        if not attr_name:
            continue

        for t in all_texts:
            if related_ability_text(t, attr_name, main_attr, sub_attr):
                ability_source_rows.append([
                    "主属性" if attr_name == main_attr else "副属性",
                    attr_name,
                    clean_text(t)
                ])

    ability_source_df = pd.DataFrame(
        ability_source_rows,
        columns=["类型", "属性", "来源词条"]
    )

    with st.expander("主 / 副属性能力值来源", expanded=False):
        if len(ability_source_df) > 0:
            st.table(ability_source_df)
        else:
            st.write("当前没有主 / 副属性相关词条。")

    base_hp = to_num(safe_get(character_data, "生命值", 0), 0)
    hp_flat = sum(stat_flat_from_text(t, ["生命值"]) for t in all_texts)
    hp_pct = sum(stat_percent_from_text(t, ["生命值"]) for t in all_texts)
    # 力量提供生命值：游戏里 250 力量 => +1250 生命值，即力量×5
    hp = floor_num((base_hp + hp_flat + strength * 5) * (1 + hp_pct))

    defense = floor_num(to_num(safe_get(character_data, "防御力", 0), 0) + sum(stat_flat_from_text(t, ["防御力"]) for t in all_texts))

    crit_rate = pct_fraction(safe_get(character_data, "暴击率", "5.0%")) + sum(panel_bonus_percent(t, ["暴击率"]) for t in all_texts)
    crit_damage = pct_fraction(safe_get(character_data, "暴击伤害", "50.0%")) + sum(panel_bonus_percent(t, ["暴击伤害"]) for t in all_texts)

    source_skill = floor_num(to_num(safe_get(character_data, "源石技艺强度", 0), 0) + sum(stat_flat_from_text(t, ["源石技艺强度", "源石技艺"]) for t in all_texts))

    physical_bonus_items = collect_physical_bonus(weapon_texts, potential_texts, equip_texts if enable_equipment else [], set_effect_texts if enable_equipment else [])
    physical_bonus = sum(v for _, _, v in physical_bonus_items)
    heat_bonus = sum(panel_bonus_percent(t, ["灼热伤害加成", "灼热伤害"]) for t in all_texts)
    electric_bonus = sum(panel_bonus_percent(t, ["电磁伤害加成", "电磁伤害"]) for t in all_texts)
    cold_bonus = sum(panel_bonus_percent(t, ["寒冷伤害加成", "寒冷伤害"]) for t in all_texts)
    nature_bonus = sum(panel_bonus_percent(t, ["自然伤害加成", "自然伤害"]) for t in all_texts)
    all_skill_bonus = sum(
        panel_bonus_percent(t, [
            "全技能伤害加成",
            "全技能伤害",
            "技能伤害加成",
            "技能伤害"
        ])
        for t in all_texts
    )

    combat_skill_bonus = all_skill_bonus + sum(
        panel_bonus_percent(t, ["战技伤害加成", "战技伤害"])
        for t in all_texts
    )

    combo_bonus = all_skill_bonus + sum(
        panel_bonus_percent(t, ["连携技伤害加成", "连携技伤害", "连携伤"])
        for t in all_texts
    )

    ultimate_bonus = all_skill_bonus + sum(
        panel_bonus_percent(t, ["终结技伤害加成", "终结技伤害"])
        for t in all_texts
    )
    charge_eff = 1 + sum(panel_bonus_percent(t, ["终结技能充能效率", "终结技充能效率"]) for t in all_texts)
    received_heal_bonus = will * 0.001 + sum(panel_bonus_percent(t, ["受治疗效率", "受治疗效率加成"]) for t in all_texts)
    heal_bonus = sum(panel_bonus_percent(t, ["治疗效率加成", "治疗效率"]) for t in all_texts)

    physical_res = attr_res_from_ability(agility)
    heat_res = attr_res_from_ability(intellect)
    electric_res = attr_res_from_ability(intellect)
    cold_res = attr_res_from_ability(intellect)
    nature_res = attr_res_from_ability(intellect)

    left_col, right_col = st.columns([1, 2])
    with left_col:
        st.markdown("### 能力值")
        show_ability("力量", strength, f"基础{to_num(base_strength):.3f} + 词条{str_flat:g}" + (f"，百分比+{str_pct*100:.1f}%" if str_pct else ""), main=(main_attr == "力量"))
        show_ability("敏捷", agility, f"基础{to_num(base_agility):.3f} + 词条{agi_flat:g}" + (f"，百分比+{agi_pct*100:.1f}%" if agi_pct else ""), main=(main_attr == "敏捷"))
        show_ability("智识", intellect, f"基础{to_num(base_intellect):.3f} + 词条{int_flat:g}" + (f"，百分比+{int_pct*100:.1f}%" if int_pct else ""), main=(main_attr == "智识"))
        show_ability("意志", will, f"基础{to_num(base_will):.3f} + 词条{will_flat:g}" + (f"，百分比+{will_pct*100:.1f}%" if will_pct else ""), main=(main_attr == "意志"))

    with right_col:
        st.markdown("### 属性详情")
        show_stat("生命值", hp)
        show_stat("攻击力", atk)
        show_stat("防御力", defense)
        show_stat("暴击率", format_percent(crit_rate))
        show_stat("暴击伤害", format_percent(crit_damage))
        show_stat("源石技艺强度", source_skill)
        show_stat("物理抗性", physical_res)
        show_stat("灼热抗性", heat_res)
        show_stat("电磁抗性", electric_res)
        show_stat("寒冷抗性", cold_res)
        show_stat("自然抗性", nature_res)
        show_stat("物理伤害加成", format_percent(physical_bonus))
        show_stat("灼热伤害加成", format_percent(heat_bonus))
        show_stat("电磁伤害加成", format_percent(electric_bonus))
        show_stat("寒冷伤害加成", format_percent(cold_bonus))
        show_stat("自然伤害加成", format_percent(nature_bonus))
        show_stat("战技伤害加成", format_percent(combat_skill_bonus))
        show_stat("连携技伤害加成", format_percent(combo_bonus))
        show_stat("终结技伤害加成", format_percent(ultimate_bonus))
        show_stat("终结技充能效率", format_percent(charge_eff))
        show_stat("治疗效率加成", format_percent(heal_bonus))
        show_stat("受治疗效率", format_percent(received_heal_bonus))

    with st.expander("计算拆解"):
        st.markdown(f"""
        <div class='small-note'>
        干员数据：DataSource / {selected_character} / Lv.{character_level}<br>
        主属性：<b>{main_attr}</b>；副属性：<b>{sub_attr}</b><br>
        当前只要装备系统未启用，就不会把装备词条放进面板，方便校准无装备数值。<br><br>
        基础攻击总 = 干员攻击力 {char_atk:g} + 武器白值 {weapon_base_atk:g} = <b>{base_atk_total:g}</b><br>
        能力值乘区 = 主属性 {main_total} × 1.5% + 副属性 {sub_total} × 0.3% = <b>{ability_bonus:.4f}</b><br>
        攻击力 = (基础攻击总 × (1 + 能力值乘区) + 固定攻击 {fixed_attack:g}) × (1 + 攻击百分比 {attack_percent:.2%}) = <b>{atk}</b>
        </div>
        """, unsafe_allow_html=True)
        st.write("武器词条：", weapon_texts)
        st.write("信赖词条：", trust_texts)
        st.write("潜能词条：", potential_texts)
        st.write("装备词条：", [clean_text(x) for x in equip_texts if clean_text(x) and clean_text(x) != "无"])
        st.write("物理伤害加成来源：", [f"{src}: {v*100:.1f}% | {txt}" for src, txt, v in physical_bonus_items])
        st.write("参与能力值/面板计算的全部文本：", all_texts)
    # =========================
    # 自动保存当前槽位最终面板：排轴页面直接读取这里
    # =========================
    if "team_panel_results" not in st.session_state:
        st.session_state.team_panel_results = {}

    current_slot_label = slot_labels[st.session_state.active_slot]

    st.session_state.team_panel_results[current_slot_label] = {
        "干员": selected_character,
        "干员属性": safe_get(character_data, "干员属性", "物理"),
        "等级": character_level,
        "潜能": potential_level,
        "信赖": trust_level,
        "攻击力": atk,
        "基础总攻击": base_total_attack,
        "主属性": main_attr,
        "副属性": sub_attr,
        "主属性值": main_total,
        "副属性值": sub_total,
        "暴击率": crit_rate,
        "暴击伤害": crit_damage,
        "物理伤害加成": physical_bonus,
        "灼热伤害加成": heat_bonus,
        "电磁伤害加成": electric_bonus,
        "寒冷伤害加成": cold_bonus,
        "自然伤害加成": nature_bonus,
        "战技伤害加成": combat_skill_bonus,
        "连携技伤害加成": combo_bonus,
        "终结技伤害加成": ultimate_bonus,
        "技能等级": {
            "普通攻击": skill_levels.get("普通攻击", "Lv9"),
            "战技": skill_levels.get("战技", "Lv9"),
            "连携技": skill_levels.get("连携技", "Lv9"),
            "终结技": skill_levels.get("终结技", "Lv9"),
        },
    }

    # =========================
    # 四人编队：本轮交互结束后自动保存当前槽位配置
    # =========================
    if "team_profiles" in st.session_state and "active_slot" in st.session_state:
        save_current_slot(st.session_state.active_slot)

elif page_mode == "伤害计算":
    st.markdown("## 伤害计算")

    SLOT_LABELS = ["1号位", "2号位", "3号位", "4号位"]
    DAMAGE_TYPES = ["物理", "灼热", "电磁", "寒冷", "自然"]

    MEDICINE_EXCEL = next(BASE_DIR.glob("zmd伤害计算器*.xlsx"), None)

    if MEDICINE_EXCEL is None:
        st.error("找不到药剂 Excel，请把 zmd伤害计算器.xlsx 放在 py 文件同一文件夹。")
        st.stop()

    @st.cache_data(show_spinner=False)
    def load_medicine_data(path: str):
        raw = pd.read_excel(path, sheet_name="药物库", header=None)

        meds = raw.iloc[3:14, 0:8].copy()
        meds.columns = [
            "buff名称", "颜色", "描述", "攻击百分比", "固定攻击力",
            "暴击率", "伤害加成", "能力值",
        ]

        meds = meds.dropna(subset=["buff名称"])
        meds = meds[meds["buff名称"].astype(str).str.strip() != ""]
        return meds

    def get_med_value(row, col):
        if row is None or col not in row.index:
            return 0.0
        return to_num(row[col], 0)

    def apply_medicine(panel, med_row):
        atk = to_num(panel.get("攻击力", 0), 0)
        base_total_attack = to_num(panel.get("基础总攻击", atk), atk)
        crit_rate = to_num(panel.get("暴击率", 0), 0)
        crit_damage = to_num(panel.get("暴击伤害", 0), 0)
        main_value = to_num(panel.get("主属性值", 0), 0)
        sub_value = to_num(panel.get("副属性值", 0), 0)

        physical = to_num(panel.get("物理伤害加成", 0), 0)
        heat = to_num(panel.get("灼热伤害加成", 0), 0)
        electric = to_num(panel.get("电磁伤害加成", 0), 0)
        cold = to_num(panel.get("寒冷伤害加成", 0), 0)
        nature = to_num(panel.get("自然伤害加成", 0), 0)

        skill_bonus_common = {
            "战技伤害加成": to_num(panel.get("战技伤害加成", 0), 0),
            "连携技伤害加成": to_num(panel.get("连携技伤害加成", 0), 0),
            "终结技伤害加成": to_num(panel.get("终结技伤害加成", 0), 0),
        }

        if med_row is None:
            return {
                "攻击力": atk,
                "暴击率": crit_rate,
                "暴击伤害": crit_damage,
                "物理伤害加成": physical,
                "灼热伤害加成": heat,
                "电磁伤害加成": electric,
                "寒冷伤害加成": cold,
                "自然伤害加成": nature,
                **skill_bonus_common,
            }

        med_desc = clean_text(med_row.get("描述", ""))
        atk_pct = get_med_value(med_row, "攻击百分比")
        fixed_atk = get_med_value(med_row, "固定攻击力")
        med_crit_rate = get_med_value(med_row, "暴击率")
        med_crit_damage = get_med_value(med_row, "暴击伤害")
        med_damage_bonus = get_med_value(med_row, "伤害加成")

        ability_flat = 0
        m = re.search(r"增加([0-9]+(?:\.[0-9]+)?)点全能力", med_desc)
        if m:
            ability_flat = float(m.group(1))

        new_main_value = main_value + ability_flat
        new_sub_value = sub_value + ability_flat
        ability_bonus = new_main_value * 0.005 + new_sub_value * 0.002

        final_base_attack = base_total_attack * (1 + atk_pct) + fixed_atk
        final_atk = math.floor(final_base_attack * (1 + ability_bonus))

        if "物理伤害" in med_desc:
            physical += med_damage_bonus
        elif "所有伤害" in med_desc:
            physical += med_damage_bonus
            heat += med_damage_bonus
            electric += med_damage_bonus
            cold += med_damage_bonus
            nature += med_damage_bonus

        return {
            "攻击力": final_atk,
            "暴击率": crit_rate + med_crit_rate,
            "暴击伤害": crit_damage + med_crit_damage,
            "物理伤害加成": physical,
            "灼热伤害加成": heat,
            "电磁伤害加成": electric,
            "寒冷伤害加成": cold,
            "自然伤害加成": nature,
            **skill_bonus_common,
        }

    st.markdown("### 当前编队")
    medicine_df = load_medicine_data(str(MEDICINE_EXCEL))
    if medicine_df.empty:
        st.warning("没有读取到药剂数据，请确认 xlsx 文件和 py 文件在同一文件夹。")

    medicine_options = ["不使用"]
    medicine_map = {}
    if not medicine_df.empty:
        for _, row in medicine_df.iterrows():
            name = str(row.get("buff名称", "")).strip()
            if name:
                medicine_options.append(name)
                medicine_map[name] = row

    rows = []
    final_panel_map = {}

    st.markdown("#### 药剂选择")
    med_cols = st.columns(4)
    selected_meds = {}
    for idx, slot in enumerate(SLOT_LABELS):
        panel = st.session_state.team_panel_results.get(slot)
        saved_med = st.session_state.team_medicine_choices.get(slot, "不使用")
        if saved_med not in medicine_options:
            saved_med = "不使用"
        with med_cols[idx]:
            if not panel:
                st.write(f"{slot} 未配置")
                selected_meds[slot] = "不使用"
                st.session_state.team_medicine_choices[slot] = "不使用"
            else:
                picked = st.selectbox(
                    f"{slot}｜{panel.get('干员', '')}",
                    medicine_options,
                    index=medicine_options.index(saved_med),
                    key=f"medicine_select_{slot}",
                )
                selected_meds[slot] = picked
                st.session_state.team_medicine_choices[slot] = picked

    for slot in SLOT_LABELS:
        panel = st.session_state.team_panel_results.get(slot)
        if not panel:
            rows.append([slot, "未配置", "不使用", "-", "-", "-", "-", "-", "-", "-", "-"])
            continue

        picked_med = selected_meds.get(slot, "不使用")
        med_row = medicine_map.get(picked_med)
        final_panel = apply_medicine(panel, med_row)
        final_panel_map[slot] = final_panel

        rows.append([
            slot,
            panel.get("干员", ""),
            picked_med,
            final_panel["攻击力"],
            f"{final_panel['暴击率'] * 100:.1f}%",
            f"{final_panel['暴击伤害'] * 100:.1f}%",
            f"{final_panel['物理伤害加成'] * 100:.1f}%",
            f"{final_panel['灼热伤害加成'] * 100:.1f}%",
            f"{final_panel['电磁伤害加成'] * 100:.1f}%",
            f"{final_panel['寒冷伤害加成'] * 100:.1f}%",
            f"{final_panel['自然伤害加成'] * 100:.1f}%",
        ])

    st.table(pd.DataFrame(
        rows,
        columns=[
            "位置", "干员", "药剂", "攻击力", "暴击率", "暴击伤害",
            "物理伤害", "灼热伤害", "电磁伤害", "寒冷伤害", "自然伤害",
        ],
    ))

    # =========================
    # 技能浏览
    # =========================
    st.markdown("### 技能浏览")
    skill_view_slots = []
    slot_display_map_for_skill = {}
    for slot in SLOT_LABELS:
        panel = st.session_state.team_panel_results.get(slot)
        if not panel:
            continue
        char_name = panel.get("干员", "未配置")
        display_text = f"{slot}｜{char_name}"
        skill_view_slots.append(display_text)
        slot_display_map_for_skill[display_text] = slot

    if skill_view_slots:
        picked_skill_slot_display = st.selectbox(
            "选择查看技能的干员",
            skill_view_slots,
            key="skill_browser_slot"
        )
        picked_skill_slot = slot_display_map_for_skill[picked_skill_slot_display]
        picked_panel = st.session_state.team_panel_results.get(picked_skill_slot, {})
        picked_char_name = picked_panel.get("干员", "")
        picked_skill_levels = picked_panel.get("技能等级", {})
        picked_skill_data_map = load_character_skill_data(str(CHARACTER_EXCEL), picked_char_name)
        skill_card_cols = st.columns(4)
        for col, skill_name in zip(skill_card_cols, ["普通攻击", "战技", "连携技", "终结技"]):
            with col:
                show_skill_card(
                    skill_name,
                    picked_skill_data_map.get(skill_name, {}),
                    picked_skill_levels.get(skill_name, "Lv9"),
                    f"damage_skill_browser_{picked_skill_slot}_{skill_name}"
                )
    else:
        st.info("当前没有已保存的编队技能。请先在编队页保存干员。")

    # =========================
    # 伤害轴：添加动作 + 累计伤害
    # =========================
    st.markdown("### 伤害轴")
    total_metric_box = st.empty()

    def calc_axis_damage(row):
        slot = row.get("位置", "")
        panel = final_panel_map.get(slot, {})
        base_panel = st.session_state.team_panel_results.get(slot, {})
        if not panel:
            return 0, {}

        action_type = row.get("动作", "普通攻击")
        damage_type = row.get("伤害属性", "物理")
        crit_mode = row.get("暴击模式", "期望")

        atk = to_num(panel.get("攻击力", 0), 0)
        atk_pct = to_num(row.get("攻击%", 0), 0) / 100
        fixed_atk = to_num(row.get("固定攻击", 0), 0)
        final_atk_for_calc = atk * (1 + atk_pct) + fixed_atk

        crit_rate = to_num(panel.get("暴击率", 0), 0) + to_num(row.get("额外暴击率%", 0), 0) / 100
        crit_damage = to_num(panel.get("暴击伤害", 0), 0) + to_num(row.get("额外暴伤%", 0), 0) / 100
        if crit_mode == "期望":
            crit_mult = 1 + crit_rate * crit_damage
        elif crit_mode == "必暴击":
            crit_mult = 1 + crit_damage
        else:
            crit_mult = 1

        base_damage_bonus = to_num(panel.get(f"{damage_type}伤害加成", 0), 0)
        extra_attr_bonus = to_num(row.get("属性增伤%", 0), 0) / 100

        skill_bonus = 0
        if action_type == "战技":
            skill_bonus = to_num(panel.get("战技伤害加成", 0), 0)
        elif action_type == "连携技":
            skill_bonus = to_num(panel.get("连携技伤害加成", 0), 0)
        elif action_type == "终结技":
            skill_bonus = to_num(panel.get("终结技伤害加成", 0), 0)
        skill_bonus += to_num(row.get("技能增伤%", 0), 0) / 100

        damage_bonus_total = base_damage_bonus + extra_attr_bonus + skill_bonus
        vuln_final = to_num(row.get("易伤%", 0), 0) / 100
        fragile_final = to_num(row.get("脆弱%", 0), 0) / 100
        amp_final = to_num(row.get("增幅%", 0), 0) / 100
        combo_final = to_num(row.get("连击增伤%", 0), 0) / 100
        stagger_mult = 1.3 if bool(row.get("失衡易伤30%", False)) else 1.0
        multiplier = to_num(row.get("倍率%", 0), 0) / 100

        damage = math.floor(
            final_atk_for_calc
            * multiplier
            * (1 + damage_bonus_total)
            * (1 + vuln_final)
            * (1 + fragile_final)
            * (1 + amp_final)
            * (1 + combo_final)
            * stagger_mult
            * crit_mult
        )
        detail = {
            "干员": base_panel.get("干员", row.get("干员", "")),
            "原攻击力": atk,
            "计算攻击力": math.floor(final_atk_for_calc),
            "伤害加成区": damage_bonus_total,
            "暴击倍率": crit_mult,
            "失衡倍率": stagger_mult,
        }
        return damage, detail

    available_slots = []
    slot_display_map = {}
    for slot in SLOT_LABELS:
        if slot in final_panel_map:
            char_name = st.session_state.team_panel_results.get(slot, {}).get("干员", "未配置")
            display_text = f"{slot}｜{char_name}"
            available_slots.append(display_text)
            slot_display_map[display_text] = slot

    left_col, right_col = st.columns([1, 2])

    with left_col:
        st.markdown("#### 添加动作")

        if not available_slots:
            st.info("当前没有可计算的干员。请先在编队页保存面板。")
        else:
            # 顶部选择区：压成一行，减少纵向长度
            select_cols = st.columns([1.35, 1.0, 1.0])
            with select_cols[0]:
                calc_slot_display = st.selectbox("干员", available_slots, key="axis_calc_slot")
            calc_slot = slot_display_map[calc_slot_display]
            calc_char_name = st.session_state.team_panel_results.get(calc_slot, {}).get("干员", "")
            with select_cols[1]:
                calc_action_type = st.selectbox("动作", ["普通攻击", "处决", "战技", "连携技", "终结技"], key="axis_calc_action_type")
            with select_cols[2]:
                calc_damage_type = st.selectbox("属性", DAMAGE_TYPES, key="axis_calc_damage_type")

            st.markdown("##### 倍率便捷计算")
            quick_cols = st.columns([1, 1, 1, 0.95])
            with quick_cols[0]:
                quick_m1 = st.number_input("倍率1%", value=0.0, step=1.0, key="quick_m1")
            with quick_cols[1]:
                quick_m2 = st.number_input("倍率2%", value=0.0, step=1.0, key="quick_m2")
            with quick_cols[2]:
                quick_m3 = st.number_input("倍率3%", value=0.0, step=1.0, key="quick_m3")

            quick_total = quick_m1 + quick_m2 + quick_m3
            with quick_cols[3]:
                st.markdown(
                    f"""
                    <div style='background:#eaf3ff;border:1px solid #d8e8ff;border-radius:10px;
                                padding:7px 8px;margin-top:25px;text-align:center;'>
                        <div style='font-size:11px;color:#6b7280;font-weight:700;'>合计</div>
                        <div style='font-size:20px;font-weight:900;color:#111827;'>{quick_total:.1f}%</div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            if "axis_calc_multiplier" not in st.session_state:
                st.session_state["axis_calc_multiplier"] = 0.0

            if st.button("使用合计倍率", use_container_width=True, key="use_quick_multiplier_btn"):
                st.session_state["axis_calc_multiplier"] = float(quick_total)
                st.rerun()

            st.markdown("##### 参数区")
            # 三列排布，单个输入框更窄，整体高度也更短
            r1 = st.columns(3)
            with r1[0]:
                calc_multiplier = st.number_input("倍率%", value=0.0, step=1.0, key="axis_calc_multiplier")
            with r1[1]:
                fixed_atk_buff = st.number_input("固定攻击", value=0.0, step=10.0, key="axis_fixed_atk_buff")
            with r1[2]:
                atk_pct_buff = st.number_input("攻击%", value=0.0, step=1.0, key="axis_atk_pct_buff")

            r2 = st.columns(3)
            with r2[0]:
                extra_crit_rate = st.number_input("额外暴击率%", value=0.0, step=1.0, key="axis_extra_crit_rate")
            with r2[1]:
                extra_crit_damage = st.number_input("额外暴伤%", value=0.0, step=1.0, key="axis_extra_crit_damage")
            with r2[2]:
                extra_damage_bonus = st.number_input("属性增伤%", value=0.0, step=1.0, key="axis_extra_damage_bonus")

            r3 = st.columns(3)
            with r3[0]:
                skill_bonus_buff = st.number_input("技能增伤%", value=0.0, step=1.0, key="axis_skill_bonus_buff")
            with r3[1]:
                vuln_buff = st.number_input("易伤%", value=0.0, step=1.0, key="axis_vuln_buff")
            with r3[2]:
                fragile_buff = st.number_input("脆弱%", value=0.0, step=1.0, key="axis_fragile_buff")

            r4 = st.columns(3)
            with r4[0]:
                amp_buff = st.number_input("增幅%", value=0.0, step=1.0, key="axis_amp_buff")
            with r4[1]:
                combo_buff = st.number_input("连击增伤%", value=0.0, step=1.0, key="axis_combo_buff")
            with r4[2]:
                calc_is_expect = st.selectbox("暴击", ["期望", "不暴击", "必暴击"], key="axis_calc_crit_mode")

            fixed_cols = st.columns([1.0, 1.25])
            with fixed_cols[0]:
                use_stagger_bonus = st.toggle("失衡易伤30%", value=False, key="axis_use_stagger_bonus")

            preview_row = {
                "位置": calc_slot,
                "干员": calc_char_name,
                "动作": calc_action_type,
                "伤害属性": calc_damage_type,
                "倍率%": calc_multiplier,
                "攻击%": atk_pct_buff,
                "固定攻击": fixed_atk_buff,
                "额外暴击率%": extra_crit_rate,
                "额外暴伤%": extra_crit_damage,
                "技能增伤%": skill_bonus_buff,
                "属性增伤%": extra_damage_bonus,
                "易伤%": vuln_buff,
                "脆弱%": fragile_buff,
                "增幅%": amp_buff,
                "连击增伤%": combo_buff,
                "失衡易伤30%": use_stagger_bonus,
                "暴击模式": calc_is_expect,
            }
            preview_damage, _ = calc_axis_damage(preview_row)
            with fixed_cols[1]:
                st.metric("本次预览伤害", f"{preview_damage:,.0f}")

            submit_action = st.button("添加到伤害轴", use_container_width=True, key="add_action_to_axis_btn")
            if submit_action:
                preview_row["本次伤害"] = preview_damage
                st.session_state.damage_axis_rows.append(preview_row.copy())
                st.session_state.damage_axis_editor_version += 1
                st.rerun()
    with right_col:
        st.markdown("#### 当前伤害轴")

        if st.session_state.damage_axis_rows:
            # 先按当前动作顺序重算每段伤害和总伤害
            working_rows = []
            final_total = 0
            for idx, raw_row in enumerate(st.session_state.damage_axis_rows, start=1):
                row = raw_row.copy()
                dmg, _ = calc_axis_damage(row)
                final_total += dmg
                row["序号"] = idx
                row["本次伤害"] = dmg
                row["累计伤害"] = final_total
                if "删除" not in row:
                    row["删除"] = False
                working_rows.append(row)

            # 总伤害放在右侧伤害轴最上面，方便一眼看累计结果
            top_m1, top_m2 = st.columns([1.35, 1.0])
            with top_m1:
                st.metric("伤害累计", f"{final_total:,.0f}")
            with top_m2:
                st.metric("动作数", len(st.session_state.damage_axis_rows))

            # 左边是动作轴表格，右边是调序按钮
            axis_table_col, axis_order_col = st.columns([3.2, 1.05])

            with axis_table_col:
                axis_df = pd.DataFrame(working_rows)

                # 主表只显示：序号 / 删除 / 角色 / 该动作伤害 / 动作类型
                # 其他攻击、易伤、暴伤、倍率等细节仍保存在后台，不会丢。
                display_cols = ["序号", "删除", "干员", "本次伤害", "动作"]
                axis_df = axis_df[[c for c in display_cols if c in axis_df.columns]]

                edited_axis_df = st.data_editor(
                    axis_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "序号": st.column_config.NumberColumn(
                            "序号",
                            width="small"
                        ),
                        "删除": st.column_config.CheckboxColumn(
                            "删除",
                            width="small"
                        ),
                        "干员": st.column_config.TextColumn(
                            "角色",
                            width="medium"
                        ),
                        "本次伤害": st.column_config.NumberColumn(
                            "该动作伤害",
                            format="%d",
                            width="medium"
                        ),
                        "动作": st.column_config.SelectboxColumn(
                            "动作类型",
                            options=["普通攻击", "处决", "战技", "连携技", "终结技"],
                            required=True,
                            width="medium"
                        ),
                    },
                    disabled=["序号", "干员", "本次伤害"],
                    key=f"damage_axis_editor_{st.session_state.damage_axis_editor_version}"
                )

                if st.button("保存表格修改 / 删除勾选项", use_container_width=True, key="save_axis_editor_btn"):
                    original_rows = list(st.session_state.damage_axis_rows)
                    updated_rows = []

                    for _, raw_row in edited_axis_df.iterrows():
                        edited = raw_row.to_dict()

                        if bool(edited.get("删除", False)):
                            continue

                        seq = int(to_num(edited.get("序号", 0), 0)) - 1
                        if 0 <= seq < len(original_rows):
                            row = original_rows[seq].copy()
                        else:
                            continue

                        # 主表只允许改动作类型；其他隐藏参数继续保留。
                        if "动作" in edited:
                            row["动作"] = edited["动作"]

                        slot = row.get("位置", "")
                        row["干员"] = st.session_state.team_panel_results.get(slot, {}).get(
                            "干员",
                            row.get("干员", "")
                        )
                        row["删除"] = False
                        updated_rows.append(row)

                    st.session_state.damage_axis_rows = updated_rows
                    st.session_state.damage_axis_editor_version += 1
                    st.rerun()

            with axis_order_col:
                st.markdown("##### 操作")
                if "confirm_clear_axis" not in st.session_state:
                    st.session_state.confirm_clear_axis = False

                if not st.session_state.confirm_clear_axis:
                    if st.button("清空伤害轴", use_container_width=True, key="clear_axis_prepare_btn"):
                        st.session_state.confirm_clear_axis = True
                        st.rerun()
                else:
                    st.warning("确认清空？")
                    clear_confirm_cols = st.columns(2)
                    with clear_confirm_cols[0]:
                        if st.button("确认", use_container_width=True, key="clear_axis_confirm_btn"):
                            st.session_state.damage_axis_rows = []
                            st.session_state.damage_axis_editor_version += 1
                            st.session_state.confirm_clear_axis = False
                            st.rerun()
                    with clear_confirm_cols[1]:
                        if st.button("取消", use_container_width=True, key="clear_axis_cancel_btn"):
                            st.session_state.confirm_clear_axis = False
                            st.rerun()

                st.markdown("---")
                st.markdown("##### 调序")
                move_index = st.selectbox(
                    "选择序号",
                    list(range(1, len(st.session_state.damage_axis_rows) + 1)),
                    key="axis_move_index"
                )

                if st.button("上移", use_container_width=True, key="axis_move_up_btn"):
                    idx = int(move_index) - 1
                    if idx > 0:
                        rows_tmp = st.session_state.damage_axis_rows
                        rows_tmp[idx - 1], rows_tmp[idx] = rows_tmp[idx], rows_tmp[idx - 1]
                        st.session_state.damage_axis_editor_version += 1
                        st.rerun()

                if st.button("下移", use_container_width=True, key="axis_move_down_btn"):
                    idx = int(move_index) - 1
                    if idx < len(st.session_state.damage_axis_rows) - 1:
                        rows_tmp = st.session_state.damage_axis_rows
                        rows_tmp[idx + 1], rows_tmp[idx] = rows_tmp[idx], rows_tmp[idx + 1]
                        st.session_state.damage_axis_editor_version += 1
                        st.rerun()

        else:
            st.info("伤害轴还是空的。先在左边添加一个动作。")
            st.metric("伤害累计", "0")

# 记录本轮所在页面，用于下次切回编队时恢复当前槽位
st.session_state["_last_page"] = page_mode
